from django.shortcuts import render,get_object_or_404
from .models import Department,EmployeeInfo,Task, TaksInfo
from .serializer import DepartmentSerializer,EmployeeSerializer,UserSerializer,TaskSerializer
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.decorators import api_view
from django.views.generic import View

# from django.template.loader import get_template
from django.template.loader import render_to_string
from weasyprint import HTML
from openpyxl import Workbook
from django.http import HttpResponse
# Create your views here.

class DepartmentListCreateView(generics.ListCreateAPIView):
    queryset=Department.objects.all()
    serializer_class=DepartmentSerializer

class DepartmentUpdateDestryView(generics.RetrieveUpdateDestroyAPIView):
    queryset=Department.objects.all()
    serializer_class=DepartmentSerializer

class EmployeeInfoListCreateView(generics.ListCreateAPIView):
    queryset = EmployeeInfo.objects.select_related('user', 'department').all()
    serializer_class = EmployeeSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        data = []
        for employee in queryset:
            data.append({
                'id': employee.id,
                'user': employee.user.username,
                'department': employee.department.name,
                'role': employee.role,
                'salary': employee.salary,

            })
        return Response(data)


class EmployeeListUpdateDestryView(generics.RetrieveUpdateDestroyAPIView):
    queryset=EmployeeInfo.objects.select_related('user','department').all()
    serializer_class=EmployeeSerializer


    
@api_view(['GET',])
def get_last_employee_task(request, pk):
    last_task_info = get_object_or_404(Task, pk=pk)
    print(last_task_info)
    if request.method == 'GET':
        if not last_task_info:
            return Response({"detail": "No tasks assigned"}, status=404)

        data = {
            "id": last_task_info.task.id,
            "task": last_task_info.task.task,
            "description": last_task_info.task.description,
            "deadline": last_task_info.task.deadline,
            "status": last_task_info.task.status,
            "assigned_at": last_task_info.assigned_at,
        }
        return Response(data)


class TaskUpdateDestryView(generics.RetrieveUpdateDestroyAPIView):
    queryset=Task.objects.all()
    serializer_class=TaskSerializer

{

# class EmployeeTaskPdfView(generics.RetrieveAPIView):
#     queryset = EmployeeInfo.objects.all()
#     template = get_template('pdf_index.html')


#     def get(self, request, *args, **kwargs):
#         employee = self.get_object()
#         tasks = Task.objects.filter(employee=employee).order_by('-assigned_at')
#         tasks_info = TaksInfo.objects.filter(employee=employee).order_by('-assigned_at')

       
#         buffer = BytesIO()
#         p = canvas.Canvas(buffer, pagesize=letter)
#         width, height = letter

#         p.drawString(100, height - 50, f"Tasks for {employee.user.username}")

#         y = height - 100
#         for taskinfo in tasks_info:
#             p.drawString(
#                 100, y,
#                 f"Task: {taskinfo.task.task}, Status: {taskinfo.task.status}, Deadline: {taskinfo.task.deadline}"
#             )
#             y -= 20
#             if y < 50:
#                 p.showPage()
#                 y = height - 50

#         p.save()
#         buffer.seek(0)

#         response = HttpResponse(buffer, content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename="tasks_{employee.user.username}.pdf"'
#         return response



# class EmployeeTaskPdfView(View):
    
#     def get(self, request, *args, **kwargs):
#         employee = EmployeeInfo.objects.get(pk=kwargs['pk'])

#         tasks_info = TaksInfo.objects.filter(employee=employee).order_by('-assigned_at')

#         buffer = BytesIO()
#         p = canvas.Canvas(buffer, pagesize=letter)
#         width, height = letter

#         p.setFont("Helvetica-Bold", 14)
#         p.drawString(50, height - 50, f"Tasks for {employee.user.username}")

#         p.setFont("Helvetica", 12)
#         p.drawString(50, height - 70, f"Role: {employee.get_role_display()}")
#         p.drawString(50, height - 90, f"Department: {employee.department.name}")
#         if employee.address:
#             p.drawString(50, height - 110, f"Address: {employee.address}")

#         y = height - 150
#         p.setFont("Helvetica-Bold", 12)
#         p.drawString(100, y, "Task")
#         p.drawString(300, y, "Status")
#         p.drawString(450, y, "Deadline")
#         y -= 30
#         line_height = 25 

#         p.setFont("Helvetica", 12)
#         if tasks_info.exists():
#             for taskinfo in tasks_info:
#                 task = taskinfo.task
#                 p.drawString(100, y, task.task)
#                 p.drawString(300, y, task.get_status_display())
#                 p.drawString(450, y, str(task.deadline) if task.deadline else "N/A")
#                 y -= line_height
#                 y -= 30
#                 if y < 50:
#                     p.showPage()
#                     y = height - 50
#         else:
#             p.drawString(100, y, "No tasks assigned")

#         p.save()
#         buffer.seek(0)

#         # response = HttpResponse(buffer, content_type='application/pdf')
#         # response['Content-Disposition'] = f'attachment; filename="tasks_{employee.user.username}.pdf"'
#         response = HttpResponse(buffer, content_type='application/pdf')
#         response['Content-Disposition'] = f'inline; filename="tasks_{employee.user.username}.pdf"'

#         return response
}



class EmployeeTaskPdfView(View):

    def get(self, request, *args, **kwargs):
        pk = self.kwargs.get('pk')
        employee = EmployeeInfo.objects.get(pk=pk)
        tasks = TaksInfo.objects.filter(employee=employee).order_by('-assigned_at')
        context = {
            'employee': employee,
            'tasks': tasks,
        }

        html_string = render_to_string('pdf_index.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tasks_{employee.user.username}.pdf"'
        return response

from openpyxl.styles import Font, PatternFill, Alignment

class EmployeeExcelView(View):
    
    def get(self, request, *args, **kwargs):
        pk=self.kwargs.get('pk')
        emp=EmployeeInfo.objects.get(pk=pk)
        tasks=TaksInfo.objects.filter(employee=emp).order_by('-assigned_at')

        wb=Workbook()
        ws=wb.active
        ws.title=f"Tasks_{emp.user.username}"

        headers = ["Task", "Status", "Deadline"]
        ws.append(headers)
        header_font= Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  
        alignment = Alignment(horizontal="center")
        for col_cell in ws[1]: 
            col_cell.font = header_font
            col_cell.fill = header_fill
            col_cell.alignment = alignment

        for taskInfo in tasks:
            ws.append(
                [taskInfo.task.task,
                taskInfo.task.status,
                taskInfo.task.deadline if taskInfo.task.deadline else "N/A"])
            
        
        rep = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        rep['Content-Disposition']=f'attachment; filename=tasks_{emp.user.username}.xlsx'
        wb.save(rep)
        return rep

